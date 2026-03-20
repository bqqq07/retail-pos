from decimal import Decimal

from django.db.models import Q
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Brand, Category, Product
from .serializers import ProductSerializer


class ProductListAPIView(APIView):
    def get(self, request):
        queryset = Product.objects.select_related('category', 'brand').prefetch_related('barcodes').filter(is_active=True)
        serializer = ProductSerializer(queryset, many=True)
        return Response(serializer.data)


class ProductSearchAPIView(APIView):
    def get(self, request):
        query = request.query_params.get('q', '').strip()
        if not query:
            return Response([], status=status.HTTP_200_OK)

        queryset = Product.objects.select_related('category', 'brand').prefetch_related('barcodes').filter(
            is_active=True,
        ).filter(Q(full_name__icontains=query) | Q(sku__icontains=query))

        serializer = ProductSerializer(queryset, many=True)
        return Response(serializer.data)


class ProductImportAPIView(APIView):
    parser_classes = (MultiPartParser,)

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

        workbook = load_workbook(filename=upload, read_only=True)
        sheet = workbook.active

        created = 0
        updated = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue

            full_name = str(row[0]).strip()
            sku = str(row[1]).strip()
            category_name = str(row[2]).strip() if len(row) > 2 and row[2] else 'General'
            brand_name = str(row[3]).strip() if len(row) > 3 and row[3] else None
            sale_price = Decimal(str(row[4])) if len(row) > 4 and row[4] is not None else Decimal('0')

            category, _ = Category.objects.get_or_create(name=category_name, defaults={'parent': None})
            brand = None
            if brand_name:
                brand, _ = Brand.objects.get_or_create(name=brand_name)

            defaults = {
                'full_name': full_name,
                'category': category,
                'brand': brand,
                'sale_price': sale_price,
            }
            product, created_flag = Product.objects.update_or_create(sku=sku, defaults=defaults)

            if created_flag:
                created += 1
            else:
                updated += 1

        return Response({'created': created, 'updated': updated}, status=status.HTTP_200_OK)
